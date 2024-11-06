# flake8: noqa
from constance import config
from django.conf import settings
from drf_yasg.utils import swagger_auto_schema
from openpyxl import load_workbook
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.file.models import File


class ExampleFileSerializer(serializers.Serializer):
    url = serializers.URLField()


class ExampleImportFileUrlBuilder:

    def __init__(self, file_constance):
        self.file_constance = file_constance

    def get_url(self, request):
        return request.build_absolute_uri(settings.MEDIA_URL + getattr(config, self.file_constance, ''))


class LoadEntitiesSerializer(serializers.Serializer):
    file = serializers.PrimaryKeyRelatedField(queryset=File.objects.all())

    def to_representation(self, instance):
        return {'file_id': instance['file'].id}


class LoadEntitiesExampleViewsetMixin:
    example_file_constance = None

    @swagger_auto_schema(responses={201: ExampleFileSerializer})
    @action(methods=['get'], detail=False)
    def example(self, request):
        url_builder = ExampleImportFileUrlBuilder(self.example_file_constance)
        return Response({'url': url_builder.get_url(request)})


class XlsxLoader:
    """
    Нужно
        - указать с какой строки начинать импорт(отсчет с нуля)
        - перегрузить process_row - импорт одной строки, без обработки ошибок и подсчета обработанных и ошибочных
        - по необходимости перегрузить prepare - вызывается один раз перед обработкой файла
    """
    start_row = 1

    def __init__(self):
        self.kwargs = None
        self.sheet = None
        self.errors = []
        self.processed = 0

    def process_row(self, row):
        raise NotImplementedError()

    def prepare(self):
        pass

    def import_file(self, file: File, **kwargs) -> dict:
        self.kwargs = kwargs

        wb = load_workbook(filename=file.file.path)
        sheet = wb.worksheets[0]

        self.sheet = sheet

        self.prepare()

        self.errors = []
        self.processed = 0

        for i, row in enumerate(sheet.iter_rows(self.start_row + 1, values_only=True)):
            try:
                self.process_row(row)
                self.processed += 1
            except Exception as e:
                self.errors.append([i + self.start_row, str(e)])

        return {'processed': self.processed, 'errors': self.errors}
